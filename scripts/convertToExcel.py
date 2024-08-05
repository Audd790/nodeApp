import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook

# Connect to MySQL database
connection = pymysql.connect(host='localhost',
                             user='root',
                             password='auddii98',
                             database='absenrajawali')

# Create a new Excel workbook
wb = Workbook()
# wb.template = True
ws = wb.active

# Execute SQL query and fetch data
query = "SELECT * FROM izinKaryawan"
cursor = connection.cursor()
cursor.execute(query)
data = cursor.fetchall()
ws.append(['id', 'alasan', 'nik', 'tgl_izin', 'durasi', 'durasi_dalam_bulan', 'myTimestamp'])
# Write data to Excel
for row_index, row_data in enumerate(data, start=1):
    for col_index, cell_data in enumerate(row_data, start=1):
        new_row_Index = row_index+1
        ws.cell(row=new_row_Index, column=col_index, value=cell_data)

try:
    wb.save('C:/Users/Operation/Desktop/Auddly/nodeApp/files/output.xlsx')
except:
    print('######################################################')
    print('The file is open, please close it before trying again!')
    print('######################################################')
finally:
    # Close the connection
    connection.close()

